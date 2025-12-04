import csv
from datetime import datetime, date
from io import BytesIO
from pathlib import Path

import joblib
import pandas as pd
import streamlit as st
from fpdf import FPDF
from src.config import MODEL_PATH  # path to models/diabetes_log_reg.pkl

from io import BytesIO
from openpyxl.styles import Font, PatternFill, Alignment

# QR code (optional)
try:
    import qrcode

    QR_AVAILABLE = True
except ImportError:
    QR_AVAILABLE = False

# ---------- PATH / CONFIG ----------
PROJECT_ROOT = Path(__file__).resolve().parent

REPORTS_DIR = PROJECT_ROOT / "reports"
REPORTS_DIR.mkdir(exist_ok=True)

HISTORY_CSV = REPORTS_DIR / "history.csv"


# ---------- MODEL LOADING ----------
@st.cache_resource
def load_model():
    """Load trained ML model."""
    return joblib.load(MODEL_PATH)


# ---------- DEFAULT VALUES ----------
DEFAULTS = {
    "pregnancies": 1,
    "glucose": 120.0,
    "blood_pressure": 70.0,
    "skin_thickness": 20.0,
    "insulin": 80.0,
    "bmi": 25.0,
    "dpf": 0.5,
    "age": 30,
}


def init_session_state():
    """Initialize session defaults."""
    for key, value in DEFAULTS.items():
        st.session_state.setdefault(key, value)

    # Patient info
    st.session_state.setdefault("patient_name", "")
    st.session_state.setdefault("patient_id", "")
    st.session_state.setdefault("gender", "Male")
    st.session_state.setdefault("doctor_name", "")
    st.session_state.setdefault("test_date", date.today())

    # NEW: clinical info
    st.session_state.setdefault("symptoms", "")
    st.session_state.setdefault("notes", "")

    # Prediction + history
    st.session_state.setdefault("last_result", None)
    st.session_state.setdefault("history", [])

    # Reset flag for "Next Patient" / "Reset All"
    st.session_state.setdefault("reset_to_zero", False)


def apply_reset_if_needed():
    """Reset fields to standard defaults when the flag is set."""
    if st.session_state.get("reset_for_next", False):
        # numeric fields: back to DEFAULT standard values
        for key, value in DEFAULTS.items():
            st.session_state[key] = value

        # patient-specific info: cleared
        st.session_state["patient_name"] = ""
        st.session_state["patient_id"] = ""
        st.session_state["gender"] = "Male"
        st.session_state["doctor_name"] = ""
        st.session_state["test_date"] = date.today()
        st.session_state["symptoms"] = ""
        st.session_state["notes"] = ""

        st.session_state["reset_for_next"] = False


def classify_risk(prob: float) -> str:
    """Simple risk categorization."""
    if prob < 0.33:
        return "Low"
    elif prob < 0.66:
        return "Moderate"
    return "High"


def log_history(record: dict):
    """Save prediction to in-memory history + CSV."""
    st.session_state["history"].append(record)

    fieldnames = [
        "timestamp",
        "patient_name",
        "patient_id",
        "gender",
        "doctor_name",
        "test_date",
        "age",
        "symptoms",
        "notes",
        "risk_level",
        "probability",
        "prediction",
        "pdf_name",
        "pregnancies",
        "glucose",
        "blood_pressure",
        "skin_thickness",
        "insulin",
        "bmi",
        "dpf",
    ]

    file_exists = HISTORY_CSV.exists()
    with HISTORY_CSV.open("a", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            writer.writeheader()
        writer.writerow(record)


# ---------- QR CODE ----------
def generate_qr_image(text: str, out_path: Path) -> Path:
    """Generate a QR code PNG with the given text."""
    if not QR_AVAILABLE:
        return None
    img = qrcode.make(text)
    img.save(out_path)
    return out_path


# ---------- PDF CREATION ----------
def create_pdf_report(input_df: pd.DataFrame,
                      prob: float,
                      pred_class: int,
                      patient_info: dict,
                      risk_level: str) -> Path:
    """Create a lab-style PDF report and return the file path."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    file_path = REPORTS_DIR / f"diabetes_report_{ts}.pdf"

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # ---- HEADER / LAB INFO ----
    pdf.set_font("Arial", "B", 18)
    pdf.cell(0, 10, "ABC Diagnostics Laboratory", ln=True, align="C")

    pdf.set_font("Arial", "", 11)
    pdf.cell(0, 6, "123 Health Street, Wellness City, India", ln=True, align="C")
    pdf.cell(0, 6, "Phone: +91-9876543210 | Email: info@abcdiagnostics.com", ln=True, align="C")
    pdf.ln(4)
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 8, "DIABETES RISK EVALUATION REPORT", ln=True, align="C")

    pdf.ln(4)
    pdf.set_font("Arial", "", 10)
    pdf.cell(0, 5, f"Report generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=True, align="R")

    # Top separator
    pdf.ln(2)
    y_line = pdf.get_y()
    pdf.line(10, y_line, 200, y_line)
    pdf.ln(4)

    # ---- QR CODE (top-right) ----
    qr_path = None
    if QR_AVAILABLE:
        qr_text = (
            f"Patient: {patient_info['name']} (ID: {patient_info['id']}) | "
            f"Risk: {risk_level} | Prob: {prob:.2f}"
        )
        qr_path = REPORTS_DIR / f"qr_{ts}.png"
        generate_qr_image(qr_text, qr_path)
        # place QR in top-right corner
        pdf.image(str(qr_path), x=165, y=25, w=30)

    # ---- PATIENT & REPORT INFO ----
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "Patient Information", ln=True)
    pdf.set_font("Arial", "", 11)

    row = input_df.iloc[0]

    # Two-column layout
    pdf.cell(100, 6, f"Patient Name: {patient_info['name']}", ln=0)
    pdf.cell(0, 6, f"Patient ID: {patient_info['id']}", ln=1)

    pdf.cell(100, 6, f"Gender: {patient_info['gender']}", ln=0)
    pdf.cell(0, 6, f"Age: {row['Age']}", ln=1)

    pdf.cell(100, 6, f"Referring Doctor: {patient_info['doctor']}", ln=0)
    pdf.cell(0, 6, f"Test Date: {patient_info['test_date']}", ln=1)

    pdf.ln(4)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "Clinical Information", ln=True)
    pdf.set_font("Arial", "", 11)

    symptoms = patient_info.get("symptoms", "").strip() or "Not provided"
    notes = patient_info.get("notes", "").strip() or "Not provided"

    pdf.multi_cell(0, 6, f"Symptoms / Presenting Complaints: {symptoms}")
    pdf.ln(2)
    pdf.multi_cell(0, 6, f"Additional Notes / History: {notes}")

    pdf.ln(4)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(6)

    # ---- SUMMARY SECTION ----
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "Diabetes Risk Summary", ln=True)
    pdf.set_font("Arial", "", 11)

    pred_text = "Diabetes (Positive)" if pred_class == 1 else "No Diabetes (Negative)"
    pdf.cell(0, 6, f"Predicted probability of diabetes: {prob:.2f}", ln=True)
    pdf.cell(0, 6, f"Model classification: {pred_text}", ln=True)
    pdf.cell(0, 6, f"Risk category: {risk_level}", ln=True)

    pdf.ln(6)
    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
    pdf.ln(6)

    # ---- DETAILED PARAMETERS TABLE ----
    pdf.set_font("Arial", "B", 12)
    pdf.cell(0, 8, "Measured Parameters", ln=True)

    pdf.set_font("Arial", "B", 11)
    pdf.cell(60, 7, "Parameter", border=1)
    pdf.cell(40, 7, "Value", border=1)
    pdf.cell(40, 7, "Units", border=1)
    pdf.cell(40, 7, "Reference Range", border=1, ln=1)

    pdf.set_font("Arial", "", 11)

    def row_add(name, value, unit, ref):
        pdf.cell(60, 7, name, border=1)
        pdf.cell(40, 7, str(value), border=1)
        pdf.cell(40, 7, unit, border=1)
        pdf.cell(40, 7, ref, border=1, ln=1)

    row_add("Pregnancies", row["Pregnancies"], "-", "0-10 (typical)")
    row_add("Glucose", row["Glucose"], "mg/dL", "70-140")
    row_add("Blood Pressure", row["BloodPressure"], "mmHg", "90-120 / 60-80")
    row_add("Skin Thickness", row["SkinThickness"], "mm", "10-40")
    row_add("Insulin", row["Insulin"], "uU/mL", "2-25 (fasting)")
    row_add("BMI", row["BMI"], "kg/m^2", "18.5-24.9")
    row_add("Diabetes Pedigree", row["DiabetesPedigreeFunction"], "-", "Family history index")
    row_add("Age", row["Age"], "years", "-")

    pdf.ln(8)
    pdf.set_font("Arial", "I", 9)
    pdf.multi_cell(
        0,
        5,
        "Disclaimer: This report is generated by a machine learning model for "
        "educational/demo purposes and is not a substitute for professional "
        "medical advice, diagnosis, or treatment.",
    )

    pdf.ln(8)
    pdf.set_font("Arial", "", 9)
    pdf.cell(0, 5, "Authorized Signatory: ____________________", ln=True)

    pdf.output(str(file_path))
    return file_path


# ---------- STREAMLIT UI ----------
def main():
    init_session_state()
    apply_reset_if_needed()

    model = load_model()

    # ----- SIDEBAR -----
    with st.sidebar:
        st.header("Actions")

        if st.session_state["history"]:
            st.subheader("Recent Predictions")
            df_hist = pd.DataFrame(st.session_state["history"])
            df_hist = df_hist.sort_values("timestamp", ascending=False).head(5)
            for _, row in df_hist.iterrows():
                st.write(
                    f"**{row['timestamp']}**  \n"
                    f"ID: {row['patient_id'] or '-'} | "
                    f"Prob: {row['probability']:.2f} | Risk: {row['risk_level']}"
                )

        if st.button("Reset All"):
            st.session_state["last_result"] = None
            st.session_state["history"] = []
            st.session_state["reset_for_next"] = True
            st.rerun()

        if not QR_AVAILABLE:
            st.info("QR code: install with `pip install qrcode[pil]` to enable QR in PDFs.")

    # ----- MAIN CONTENT -----
    st.title("Diabetes Prediction App")
    st.write("This app uses a Logistic Regression model to predict the probability of diabetes.")
    st.write("**Note:** This is a demo tool and not a medical device.")

    # Patient info
    st.header("Patient Information")
    col1, col2 = st.columns(2)

    with col1:
        patient_name = st.text_input("Patient Name", key="patient_name")
        patient_id = st.text_input("Patient ID", key="patient_id")
        gender = st.selectbox("Gender", ["Male", "Female", "Other"], key="gender")

    with col2:
        doctor_name = st.text_input("Referring Doctor", key="doctor_name")
        test_date = st.date_input("Test Date", key="test_date")

    # Clinical info (NEW)
    st.subheader("Clinical Details")
    symptoms = st.text_area(
        "Symptoms / Presenting Complaints",
        key="symptoms",
        placeholder="e.g., frequent urination, increased thirst, fatigue...",
    )
    notes = st.text_area(
        "Additional Notes / History",
        key="notes",
        placeholder="e.g., family history, medications, lifestyle, etc.",
    )

    # Measurements
    st.header("Enter Clinical Measurements")
    colm1, colm2 = st.columns(2)

    with colm1:
        pregnancies = st.number_input("Pregnancies", 0, 20, key="pregnancies")
        glucose = st.number_input("Glucose (mg/dL)", 0.0, 300.0, key="glucose")
        blood_pressure = st.number_input("Blood Pressure (mmHg)", 0.0, 200.0, key="blood_pressure")
        skin_thickness = st.number_input("Skin Thickness (mm)", 0.0, 100.0, key="skin_thickness")

    with colm2:
        insulin = st.number_input("Insulin (uU/mL)", 0.0, 900.0, key="insulin")
        bmi = st.number_input("BMI (kg/m^2)", 0.0, 70.0, key="bmi")
        dpf = st.number_input("Diabetes Pedigree Function", 0.0, 3.0, key="dpf")
        age = st.number_input("Age (years)", 0, 120, key="age")

    input_df = pd.DataFrame([{
        "Pregnancies": pregnancies,
        "Glucose": glucose,
        "BloodPressure": blood_pressure,
        "SkinThickness": skin_thickness,
        "Insulin": insulin,
        "BMI": bmi,
        "DiabetesPedigreeFunction": dpf,
        "Age": age,
    }])

    patient_info = {
        "name": patient_name,
        "id": patient_id,
        "gender": gender,
        "doctor": doctor_name,
        "test_date": test_date.strftime("%Y-%m-%d") if isinstance(test_date, date) else str(test_date),
        "symptoms": symptoms,
        "notes": notes,
    }

    # ----- PREDICT -----
    if st.button("Predict"):
        prob = float(model.predict_proba(input_df)[0][1])
        pred_class = int(model.predict(input_df)[0])
        risk_level = classify_risk(prob)

        pdf_path = create_pdf_report(input_df, prob, pred_class, patient_info, risk_level)
        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        record = {
            "timestamp": now,
            "patient_name": patient_name,
            "patient_id": patient_id,
            "gender": gender,
            "doctor_name": doctor_name,
            "test_date": patient_info["test_date"],
            "age": age,
            "symptoms": symptoms,
            "notes": notes,
            "risk_level": risk_level,
            "probability": prob,
            "prediction": pred_class,
            "pdf_name": pdf_path.name,
            "pregnancies": pregnancies,
            "glucose": glucose,
            "blood_pressure": blood_pressure,
            "skin_thickness": skin_thickness,
            "insulin": insulin,
            "bmi": bmi,
            "dpf": dpf,
        }

        log_history(record)

        st.session_state["last_result"] = {
            "prob": prob,
            "pred_class": pred_class,
            "risk_level": risk_level,
            "pdf_bytes": pdf_bytes,
            "pdf_name": pdf_path.name,
            "patient_info": patient_info,
        }

    # ----- SHOW RESULT -----
    if st.session_state["last_result"] is not None:
        res = st.session_state["last_result"]
        st.subheader("Results")

        st.write(f"**Predicted probability of diabetes:** {res['prob']:.2f}")
        st.write(f"**Risk category:** {res['risk_level']}")

        if res["pred_class"] == 1:
            st.error("Model classification: **Diabetes (Positive)**")
        else:
            st.success("Model classification: **No Diabetes (Negative)**")

        st.download_button(
            "Download PDF report",
            data=res["pdf_bytes"],
            file_name=res["pdf_name"],
            mime="application/pdf",
        )

        if st.button("Next Patient"):
            st.session_state["last_result"] = None
            st.session_state["reset_for_next"] = True
            st.rerun()

    # ----- HISTORY & EXCEL EXPORT -----
    st.header("Patient History & Export")

    if st.session_state["history"]:
        df_hist = pd.DataFrame(st.session_state["history"])

        st.subheader("All Predictions (current session + CSV log)")
        st.dataframe(df_hist)

        # ---------- Export to nicely formatted Excel ----------
        st.subheader("Export History")

        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
            # Write raw data
            df_hist.to_excel(writer, sheet_name="Predictions", index=False)

            wb = writer.book
            ws = writer.sheets["Predictions"]

            # Freeze header row
            ws.freeze_panes = "A2"

            # Auto-filter on header row
            ws.auto_filter.ref = ws.dimensions

            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Auto-fit-ish column widths based on max length
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        cell_value = str(cell.value) if cell.value is not None else ""
                        max_length = max(max_length, len(cell_value))
                    except Exception:
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[col_letter].width = min(adjusted_width, 40)

            # Risk-level conditional coloring
            # (Assumes 'risk_level' is one of the columns)
            if "risk_level" in df_hist.columns:
                risk_col_idx = df_hist.columns.get_loc("risk_level") + 1  # 1-based
                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=risk_col_idx)
                    value = str(cell.value).strip().lower()
                    if value == "low":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "moderate":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == "high":
                        cell.fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

        excel_buffer.seek(0)

        st.download_button(
            "Download history as Excel",
            data=excel_buffer,
            file_name="diabetes_history.xlsx",
            mime=(
                "application/vnd.openxmlformats-"
                "officedocument.spreadsheetml.sheet"
            ),
        )

    else:
        st.info("No predictions yet. Run at least one prediction to see history.")


if __name__ == "__main__":
    main()
